import csv
import io
import logging
import zipfile
from datetime import datetime, timezone
from typing import Iterator
from zoneinfo import ZoneInfo

import openpyxl
import requests

from grid_client import GridClient, PriceRecord
from ercot_api import ErcotClient as ErcotAPIClient

logger = logging.getLogger(__name__)

# ERCOT timestamps are published in Central Time
_ct = ZoneInfo("America/Chicago")

_SP_LIST_DOC_URL = "https://www.ercot.com/misapp/servlets/IceDocListJsonWS?reportTypeId=10008"
_SP_LIST_DOWNLOAD_URL = "https://www.ercot.com/misdownload/servlets/mirDownload?doclookupId={doc_id}"
_EIA860_URL = "https://www.eia.gov/electricity/data/eia860/xls/eia8602024.zip"

# Column names in the NP4-160-SG Settlement_Points CSV
_COL_BUS = "ELECTRICAL_BUS"
_COL_ZONE = "SETTLEMENT_LOAD_ZONE"
_COL_PSSE = "PSSE_BUS_NAME"


def _fetch_np4_rows() -> list[dict]:
    """Fetch the latest NP4-160-SG file and return all rows."""
    resp = requests.get(_SP_LIST_DOC_URL, timeout=30)
    resp.raise_for_status()
    docs = resp.json().get("ListDocsByRptTypeRes", {}).get("DocumentList", [])
    if not docs:
        raise RuntimeError("No documents found in NP4-160-SG listing")

    latest_doc_id = docs[0]["Document"]["DocID"]
    logger.info(f"Fetching NP4-160-SG doc ID: {latest_doc_id}")

    resp = requests.get(_SP_LIST_DOWNLOAD_URL.format(doc_id=latest_doc_id), timeout=60)
    resp.raise_for_status()

    with zipfile.ZipFile(io.BytesIO(resp.content)) as zf:
        csv_name = next(n for n in zf.namelist() if "Settlement_Points" in n and n.endswith(".csv"))
        with zf.open(csv_name) as f:
            return list(csv.DictReader(io.TextIOWrapper(f, encoding="utf-8")))


def _fetch_eia860_psse_to_latlon() -> dict[str, tuple[float, float]]:
    """Download EIA Form 860 and return a psse_node_name -> (lat, lon) mapping."""
    logger.info("Downloading EIA Form 860...")
    resp = requests.get(_EIA860_URL, timeout=120)
    resp.raise_for_status()

    plant_latlon: dict[int, tuple[float, float]] = {}
    psse_to_latlon: dict[str, tuple[float, float]] = {}

    with zipfile.ZipFile(io.BytesIO(resp.content)) as zf:
        with zf.open("2___Plant_Y2024.xlsx") as f:
            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
            for i, row in enumerate(wb.active.iter_rows(values_only=True)):
                if i < 2:
                    continue
                plant_code, lat, lon, ba_code = row[2], row[9], row[10], row[12]
                if plant_code and lat and lon and ba_code == "ERCO":
                    plant_latlon[plant_code] = (float(lat), float(lon))

        with zf.open("3_1_Generator_Y2024.xlsx") as f:
            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
            for i, row in enumerate(wb.active.iter_rows(values_only=True)):
                if i < 2:
                    continue
                plant_code, node_name = row[2], row[13]
                if node_name and plant_code and plant_code in plant_latlon:
                    psse_to_latlon[str(node_name).strip()] = plant_latlon[plant_code]

    logger.info(f"EIA 860: {len(psse_to_latlon)} ERCOT resource nodes with lat/lon")
    return psse_to_latlon


class ERCOTClient(GridClient):
    def __init__(self):
        self._api = ErcotAPIClient()
        try:
            self._psse_to_latlon = _fetch_eia860_psse_to_latlon()
        except Exception as e:
            logger.warning(f"Failed to load EIA 860 geocoding data: {e}")
            self._psse_to_latlon = {}

    def grid(self) -> str:
        return "ERCOT"

    def node_type(self) -> str:
        return "ELECTRICAL_BUS"

    def initial_locations(self) -> list[dict]:
        try:
            rows = _fetch_np4_rows()
        except Exception as e:
            logger.warning(f"Failed to fetch NP4-160-SG: {e}")
            return []

        locations = []
        geocoded = 0
        for row in rows:
            bus = row.get(_COL_BUS, "").strip()
            zone = row.get(_COL_ZONE, "").strip()
            psse = row.get(_COL_PSSE, "").strip()
            if not bus:
                continue
            latlon = self._psse_to_latlon.get(psse) if psse else None
            if latlon:
                geocoded += 1
            locations.append({
                "grid": "ERCOT",
                "node_name": bus,
                "node_type": self.node_type(),
                "settlement_load_zone": zone or None,
                "latitude": latlon[0] if latlon else None,
                "longitude": latlon[1] if latlon else None,
            })

        logger.info(f"ERCOT initial_locations: {len(locations)} buses, {geocoded} geocoded")
        return locations

    def iter_pages(self, start: datetime, end: datetime) -> Iterator[list[PriceRecord]]:
        for data in self._api.iter_pages(start, end):
            meta = data["_meta"]
            logger.info(
                f"Page {meta['currentPage']}/{meta['totalPages']}, "
                f"records: {meta['totalRecords']}"
            )
            rows = data.get("data", [])
            yield [
                PriceRecord(
                    node_name=row[2],
                    timestamp_utc=datetime.fromisoformat(row[0])
                        .replace(tzinfo=_ct)
                        .astimezone(timezone.utc),
                    lmp=row[3],
                )
                for row in rows
            ]
